"""
엑셀 파서 모듈 — .xlsx 파일에서 라벨 데이터를 추출합니다.

지원 형식:
  - B열 = 필드명, C열 = 값 (라벨1.xlsx 형식)
  - 시트 내 B3="내용" 헤더 이후 B4~B9에 필드명, C4~C9에 값
"""

import os
from typing import Dict, List, Optional
from openpyxl import load_workbook


# 엑셀 필드명 → 내부 키 매핑
FIELD_MAP = {
    '제목': 'title',
    '생산연도': 'productionYear',
    '부서명': 'departmentName',
    '분류번호': 'classificationCode',
    '보존기간': 'retentionPeriod',
    '관리번호': 'managementNumber',
}

# 보존기간 유효 값
VALID_RETENTION_PERIODS = {'영구', '준영구', '30년', '10년', '5년', '3년', '1년'}


def parse_excel(file_path: str) -> List[Dict[str, str]]:
    """
    엑셀 파일을 파싱하여 라벨 데이터 리스트를 반환합니다.
    
    Args:
        file_path: .xlsx 파일 경로
        
    Returns:
        라벨 데이터 딕셔너리의 리스트
        예: [{'title': '신입사원채용공고', 'productionYear': '2026년', ...}]
        
    Raises:
        FileNotFoundError: 파일이 존재하지 않는 경우
        ValueError: 지원하지 않는 파일 형식인 경우
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {file_path}")
    
    if not file_path.lower().endswith('.xlsx'):
        raise ValueError(f"지원하지 않는 파일 형식입니다. .xlsx 파일만 지원합니다: {file_path}")
    
    wb = load_workbook(file_path, read_only=True, data_only=True)
    labels = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        label_data = _parse_sheet(ws)
        if label_data:
            labels.append(label_data)
    
    wb.close()
    
    if not labels:
        raise ValueError("엑셀 파일에서 라벨 데이터를 찾을 수 없습니다.")
    
    return labels


def _parse_sheet(ws) -> Optional[Dict[str, str]]:
    """
    시트 하나를 파싱하여 라벨 데이터를 추출합니다.
    B열에서 필드명을 찾고, 같은 행의 C열에서 값을 가져옵니다.
    """
    label = {}
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=3):
        field_cell = row[0]  # B열
        value_cell = row[1]  # C열
        
        if field_cell.value is None:
            continue
        
        field_name = str(field_cell.value).strip()
        
        if field_name in FIELD_MAP:
            raw_value = str(value_cell.value).strip() if value_cell.value else ''
            key = FIELD_MAP[field_name]
            label[key] = _normalize_value(key, raw_value)
    
    # 최소 1개 필드가 있어야 유효한 라벨
    if len(label) >= 1:
        return label
    return None


def _normalize_value(key: str, value: str) -> str:
    """
    필드별로 값을 정규화합니다.
    """
    if key == 'departmentName':
        # 엑셀의 \n(리터럴 문자열)을 실제 줄바꿈으로 변환
        value = value.replace('\\n', '\n')
    
    elif key == 'retentionPeriod':
        # 유효한 보존기간인지 확인
        if value not in VALID_RETENTION_PERIODS:
            # 부분 매칭 시도 (예: "1" → "1년")
            for valid in VALID_RETENTION_PERIODS:
                if value in valid or valid in value:
                    value = valid
                    break
    
    return value


def format_for_display(label_data: Dict[str, str]) -> List[tuple]:
    """
    라벨 데이터를 GUI 표시용 (필드명, 값) 튜플 리스트로 변환합니다.
    """
    display_names = {
        'title': '제목',
        'productionYear': '생산연도',
        'departmentName': '부서명',
        'classificationCode': '분류번호',
        'retentionPeriod': '보존기간',
        'managementNumber': '관리번호',
    }
    
    result = []
    for key in ['title', 'productionYear', 'departmentName', 
                'classificationCode', 'retentionPeriod', 'managementNumber']:
        if key in label_data:
            display_value = label_data[key].replace('\n', ' / ')  # 줄바꿈을 슬래시로 표시
            result.append((display_names[key], display_value))
    
    return result


# 단독 실행 시 테스트
if __name__ == '__main__':
    import sys
    
    test_file = sys.argv[1] if len(sys.argv) > 1 else '라벨1.xlsx'
    
    try:
        labels = parse_excel(test_file)
        for i, label in enumerate(labels):
            print(f"\n=== 라벨 {i + 1} ===")
            for field_name, value in format_for_display(label):
                print(f"  {field_name}: {value}")
            print(f"\n  [원본 데이터] {label}")
    except Exception as e:
        print(f"❌ 오류: {e}")

